from __future__ import annotations

import argparse
import json
import os
import shutil
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


SHEET_CONTACT_TYPE = {
    "Companys HR email": "General Company HR",
    "Personal HR email": "Personal HR Contact",
}


def norm(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def email_norm(value: object) -> str:
    return norm(value).lower()


def uen_norm(value: object) -> str:
    return norm(value).upper()


def run_m365(args: list[str]) -> str:
    m365_cmd = shutil.which("m365")
    m365_cmd_file = shutil.which("m365.cmd")
    m365_ps1 = shutil.which("m365.ps1")

    if not m365_cmd and not m365_cmd_file and not m365_ps1:
        appdata = os.environ.get("APPDATA", "")
        if appdata:
            npm_dir = Path(appdata) / "npm"
            candidate_cmd = npm_dir / "m365.cmd"
            candidate_ps1 = npm_dir / "m365.ps1"
            if candidate_cmd.exists():
                m365_cmd_file = str(candidate_cmd)
            elif candidate_ps1.exists():
                m365_ps1 = str(candidate_ps1)

    if m365_cmd_file:
        command = [m365_cmd_file, *args]
    elif m365_ps1:
        command = ["pwsh", "-NoProfile", "-File", m365_ps1, *args]
    elif m365_cmd:
        command = [m365_cmd, *args]
    else:
        raise RuntimeError("Unable to locate the Microsoft 365 CLI executable.")

    proc = subprocess.run(
        command,
        capture_output=True,
        text=True,
        check=False,
    )
    if proc.returncode != 0:
        raise RuntimeError(
            f"m365 command failed ({proc.returncode}): {' '.join(args)}\n{proc.stderr.strip()}"
        )
    return proc.stdout


def read_rows(workbook_path: Path) -> list[dict[str, object]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict[str, object]] = []

    for sheet_name, contact_type in SHEET_CONTACT_TYPE.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        iterator = ws.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration:
            continue

        headers = [norm(cell) for cell in header_row]
        header_index = {header: idx for idx, header in enumerate(headers) if header}

        email_header = "HR/General Email for Reference Checks"
        for row in iterator:
            email = norm(row[header_index[email_header]]) if email_header in header_index else ""
            normalized = email_norm(email)
            if not normalized:
                continue

            company_name = norm(row[header_index["Company Name"]]) if "Company Name" in header_index else ""
            company_address = norm(row[header_index["Company Address"]]) if "Company Address" in header_index else ""
            tel_contact = norm(row[header_index["Tel Contact"]]) if "Tel Contact" in header_index else ""
            company_uen = norm(row[header_index["Company UEN"]]) if "Company UEN" in header_index else ""
            company_uen_normalized = uen_norm(company_uen)
            notes = norm(row[header_index["Notes"]]) if "Notes" in header_index else ""

            rows.append(
                {
                    "Title": f"{company_name} - {email}"[:255],
                    "CompanyName": company_name,
                    "CompanyAddress": company_address,
                    "TelContact": tel_contact,
                    "HRReferenceEmail": email,
                    "HRReferenceEmailNormalized": normalized,
                    "CompanyUEN": company_uen,
                    "CompanyUENNormalized": company_uen_normalized,
                    "ContactType": contact_type,
                    "ContactPersonName": "",
                    "Notes": notes,
                    "IsActive": True,
                    "IsVerified": True,
                    "SourceSheet": sheet_name,
                }
            )

    return rows


def get_existing_item_id(web_url: str, list_title: str, normalized_email: str) -> int | None:
    escaped_email = normalized_email.replace("'", "''")
    filter_query = f"HRReferenceEmailNormalized eq '{escaped_email}'"
    raw = run_m365(
        [
            "spo",
            "listitem",
            "list",
            "--webUrl",
            web_url,
            "--listTitle",
            list_title,
            "--filter",
            filter_query,
            "--output",
            "json",
        ]
    )
    items = json.loads(raw or "[]")
    if not items:
        return None
    return int(items[0]["Id"])


def build_field_args(row: dict[str, object]) -> list[str]:
    args: list[str] = []
    for key in [
        "Title",
        "CompanyName",
        "CompanyAddress",
        "TelContact",
        "HRReferenceEmail",
        "HRReferenceEmailNormalized",
        "CompanyUEN",
        "CompanyUENNormalized",
        "ContactType",
        "ContactPersonName",
        "Notes",
        "SourceSheet",
    ]:
        args.extend([f"--{key}", str(row[key])])

    args.extend(["--IsActive", "true" if row["IsActive"] else "false"])
    args.extend(["--IsVerified", "true" if row["IsVerified"] else "false"])
    return args


def upsert_rows(web_url: str, list_title: str, rows: list[dict[str, object]]) -> tuple[int, int]:
    created = 0
    updated = 0

    for row in rows:
        existing_id = get_existing_item_id(web_url, list_title, str(row["HRReferenceEmailNormalized"]))
        if existing_id is None:
            run_m365(
                [
                    "spo",
                    "listitem",
                    "add",
                    "--webUrl",
                    web_url,
                    "--listTitle",
                    list_title,
                    *build_field_args(row),
                    "--output",
                    "json",
                ]
            )
            created += 1
        else:
            run_m365(
                [
                    "spo",
                    "listitem",
                    "set",
                    "--webUrl",
                    web_url,
                    "--listTitle",
                    list_title,
                    "--id",
                    str(existing_id),
                    *build_field_args(row),
                    "--output",
                    "json",
                ]
            )
            updated += 1

    return created, updated


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True)
    parser.add_argument("--web-url", required=True)
    parser.add_argument("--list-title", default="Approved HR Reference Contacts")
    parser.add_argument("--json-out", default="")
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    if not workbook_path.exists():
        raise FileNotFoundError(workbook_path)

    rows = read_rows(workbook_path)
    payload = {
        "workbook": str(workbook_path),
        "rows": rows,
        "count": len(rows),
    }

    if args.json_out:
        out_path = Path(args.json_out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")

    if args.apply:
        created, updated = upsert_rows(args.web_url, args.list_title, rows)
        payload["created"] = created
        payload["updated"] = updated

    print(json.dumps(payload, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
